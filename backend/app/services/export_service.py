"""Export service for generating JSON and XLSX registries."""

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.models.paper import Paper, PaperAuthor, PaperSource
from app.models.topic import PaperTopic, Topic

logger = logging.getLogger(__name__)

# Style constants for XLSX
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="e8e8f0")
CELL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="2a2a3e"),
    right=Side(style="thin", color="2a2a3e"),
    top=Side(style="thin", color="2a2a3e"),
    bottom=Side(style="thin", color="2a2a3e"),
)


class ExportService:
    """Generates JSON and XLSX exports of the paper registry."""

    def __init__(self):
        self.output_dir = settings.registry_dir

    async def export_json(self, db: AsyncSession) -> Path:
        """Export all papers as a JSON file."""
        papers = await self._load_all_papers(db)
        data = {
            "metadata": {
                "generated_at": datetime.utcnow().isoformat(),
                "total_papers": len(papers),
                "generator": "FL-Research-Monitor",
                "version": "0.1.0",
            },
            "papers": papers,
        }

        filepath = self.output_dir / "fl_research_registry.json"
        filepath.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str))
        logger.info(f"Exported {len(papers)} papers to JSON: {filepath}")
        return filepath

    async def export_xlsx(self, db: AsyncSession) -> Path:
        """Export papers as a multi-sheet XLSX workbook."""
        papers = await self._load_all_papers(db)
        topics = await self._load_topics(db)
        sources_stats = await self._load_source_stats(db)

        wb = Workbook()

        # Sheet 1: All Papers
        self._write_all_papers_sheet(wb, papers)

        # Sheet 2: By Topic
        self._write_by_topic_sheet(wb, papers, topics)

        # Sheet 3: By Source
        self._write_by_source_sheet(wb, papers)

        # Sheet 4: Statistics
        self._write_statistics_sheet(wb, papers, sources_stats)

        filepath = self.output_dir / "fl_research_registry.xlsx"
        wb.save(str(filepath))
        logger.info(f"Exported {len(papers)} papers to XLSX: {filepath}")
        return filepath

    async def _load_all_papers(self, db: AsyncSession) -> list[dict]:
        """Load all papers with relationships as dicts."""
        query = (
            select(Paper)
            .options(
                selectinload(Paper.sources),
                selectinload(Paper.authors).selectinload(PaperAuthor.author),
                selectinload(Paper.topics).selectinload(PaperTopic.topic),
            )
            .order_by(Paper.publication_date.desc())
        )
        result = await db.execute(query)
        papers = result.unique().scalars().all()

        return [
            {
                "id": p.id,
                "doi": p.doi,
                "title": p.title,
                "abstract": (p.abstract or "")[:500],
                "publication_date": p.publication_date,
                "journal": p.journal,
                "authors": "; ".join(
                    pa.author.name
                    for pa in sorted(p.authors, key=lambda x: x.position)
                    if pa.author
                ),
                "paper_type": p.paper_type,
                "open_access": p.open_access,
                "citation_count": p.citation_count,
                "sources": ", ".join(s.source_name for s in p.sources),
                "topics": ", ".join(
                    pt.topic.name for pt in p.topics if pt.topic
                ),
                "pdf_url": p.pdf_url,
                "has_pdf": p.pdf_local_path is not None,
                "validated": p.validated,
                "external_ids": p.external_ids,
                "created_at": str(p.created_at),
            }
            for p in papers
        ]

    async def _load_topics(self, db: AsyncSession) -> list[dict]:
        result = await db.execute(select(Topic))
        return [{"id": t.id, "name": t.name} for t in result.scalars().all()]

    async def _load_source_stats(self, db: AsyncSession) -> list[dict]:
        rows = (
            await db.execute(
                select(PaperSource.source_name, func.count(PaperSource.id))
                .group_by(PaperSource.source_name)
            )
        ).all()
        return [{"source": name, "count": count} for name, count in rows]

    def _write_all_papers_sheet(self, wb: Workbook, papers: list[dict]):
        """Write the AllPapers sheet."""
        ws = wb.active
        ws.title = "All Papers"

        headers = [
            "ID", "DOI", "Title", "Authors", "Date", "Journal",
            "Type", "Open Access", "Citations", "Sources", "Topics",
            "PDF URL", "Has PDF", "Validated",
        ]
        self._write_header_row(ws, headers)

        for i, p in enumerate(papers, start=2):
            ws.cell(row=i, column=1, value=p["id"])
            ws.cell(row=i, column=2, value=p["doi"])
            ws.cell(row=i, column=3, value=p["title"])
            ws.cell(row=i, column=4, value=p["authors"])
            ws.cell(row=i, column=5, value=p["publication_date"])
            ws.cell(row=i, column=6, value=p["journal"])
            ws.cell(row=i, column=7, value=p["paper_type"])
            ws.cell(row=i, column=8, value="Yes" if p["open_access"] else "No")
            ws.cell(row=i, column=9, value=p["citation_count"])
            ws.cell(row=i, column=10, value=p["sources"])
            ws.cell(row=i, column=11, value=p["topics"])
            ws.cell(row=i, column=12, value=p["pdf_url"])
            ws.cell(row=i, column=13, value="Yes" if p["has_pdf"] else "No")
            ws.cell(row=i, column=14, value="Yes" if p["validated"] else "No")

        self._auto_column_width(ws, len(headers))

    def _write_by_topic_sheet(self, wb: Workbook, papers: list[dict], topics: list[dict]):
        """Write a sheet with papers grouped by topic."""
        ws = wb.create_sheet("By Topic")
        row = 1

        for topic in topics:
            # Topic header
            ws.cell(row=row, column=1, value=topic["name"]).font = Font(
                bold=True, size=12, color="6366f1"
            )
            row += 1

            headers = ["Title", "Authors", "Date", "Journal", "DOI"]
            self._write_header_row(ws, headers, start_row=row)
            row += 1

            topic_papers = [p for p in papers if topic["name"] in p["topics"]]
            for p in topic_papers:
                ws.cell(row=row, column=1, value=p["title"])
                ws.cell(row=row, column=2, value=p["authors"])
                ws.cell(row=row, column=3, value=p["publication_date"])
                ws.cell(row=row, column=4, value=p["journal"])
                ws.cell(row=row, column=5, value=p["doi"])
                row += 1

            row += 1  # Blank row between topics

        self._auto_column_width(ws, 5)

    def _write_by_source_sheet(self, wb: Workbook, papers: list[dict]):
        """Write a sheet with papers grouped by source."""
        ws = wb.create_sheet("By Source")
        row = 1

        sources = sorted(set(
            src.strip()
            for p in papers
            for src in p["sources"].split(",")
            if src.strip()
        ))

        for source in sources:
            ws.cell(row=row, column=1, value=source.upper()).font = Font(
                bold=True, size=12, color="10b981"
            )
            row += 1

            headers = ["Title", "Date", "DOI", "Citations"]
            self._write_header_row(ws, headers, start_row=row)
            row += 1

            source_papers = [p for p in papers if source in p["sources"]]
            for p in source_papers:
                ws.cell(row=row, column=1, value=p["title"])
                ws.cell(row=row, column=2, value=p["publication_date"])
                ws.cell(row=row, column=3, value=p["doi"])
                ws.cell(row=row, column=4, value=p["citation_count"])
                row += 1

            row += 1

        self._auto_column_width(ws, 4)

    def _write_statistics_sheet(self, wb: Workbook, papers: list[dict], sources_stats: list[dict]):
        """Write summary statistics sheet."""
        ws = wb.create_sheet("Statistics")

        ws.cell(row=1, column=1, value="FL Research Monitor — Statistics").font = Font(
            bold=True, size=14, color="6366f1"
        )
        ws.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")

        row = 4
        ws.cell(row=row, column=1, value="Overview").font = Font(bold=True, size=12)
        row += 1
        stats = [
            ("Total Papers", len(papers)),
            ("With PDF", sum(1 for p in papers if p["has_pdf"])),
            ("Open Access", sum(1 for p in papers if p["open_access"])),
            ("Validated", sum(1 for p in papers if p["validated"])),
            ("Total Citations", sum(p["citation_count"] for p in papers)),
        ]
        for label, value in stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="By Source").font = Font(bold=True, size=12)
        row += 1
        for s in sources_stats:
            ws.cell(row=row, column=1, value=s["source"])
            ws.cell(row=row, column=2, value=s["count"])
            row += 1

        self._auto_column_width(ws, 2)

    @staticmethod
    def _write_header_row(ws, headers: list[str], start_row: int = 1):
        """Write a styled header row."""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_column_width(ws, num_cols: int, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for col in range(1, num_cols + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)
